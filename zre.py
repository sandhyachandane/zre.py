import psycopg2
from openpyxl import Workbook, load_workbook

def get_reference_code():
  return input("Enter the reference code: ")

conn = psycopg2.connect(
     host='borosilcp.postgres.database.azure.com',
     database='easyecom',
     user='db_user',
     password='Hello123!'
)
cursor = conn.cursor()

reference_code = get_reference_code()

sql_query = '''select
	'ZRE' as "Document Type",
	sap.invoice_no as "Sales Order Id",
	suborder->>'sku' as "Material",
	suborder->>'returned_item_quantity' as "Quantity",
	ms.accounting_unit as "Unit of Measure",
	ee.return_date as "Return Date",
	sap.plant as "Plant to",
	'RMBO' as "Storage Location to",
	453 as "Type",
	concat(reference_code, '_', credit_note_id) as "Order Id",
	ee.return_date as "Return Date",
	ee.credit_note_number as "CreditNote_Num"
from all_returns ee
cross join lateral jsonb_array_elements(ee.items) suborder
left join zecomm_new sap on
	(case when "warehouseId" <> 96778 then concat(ee.reference_code, '_', ee.invoice_number)
	else concat(ee.reference_code, '_', suborder->>'suborder_num') end) = sap.fba_ord_no and
	suborder->>'sku' = sap.material_code
left join master_sku ms on suborder->>'sku' = ms.sku
WHERE reference_code IN ('1623692', '1623737') '''

cursor.execute(sql_query, (reference_code,))
data = cursor.fetchall()


workbook = Workbook()
worksheet = workbook.active
headers = [desc[0] for desc in cursor.description]
worksheet.append(headers)

for row in data:
  worksheet.append(row)

excel_filename = f"data_for_reference_code{reference_code}.xlsx"
workbook.save(excel_filename)
print(f"Data for reference code '{reference_code}' has been created successfully.")
workbook = load_workbook(excel_filename)
worksheet = workbook.active

selected_data = []
for row in worksheet.iter_rows(min_row = 2, values_only = True):
  if row[9].startswith(reference_code):
   selected_data.append(row)

if selected_data:
  print("Selected reference code data: ")
  for row in selected_data:
    print(row)
else:
  print("Reference code not found")
cursor.close()
conn.close()
